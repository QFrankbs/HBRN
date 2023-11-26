import torch
import torch.nn as nn
import time
import numpy as np
import os
import openpyxl
from utils.pred_func import *
import gc 
def train(net, train_loader, eval_loader, args):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Epoch and Accuracy"
    sheet['A1'] = "Epoch"
    sheet['B1'] = "train_Accuracy"
    sheet['C1'] = "train_Loss"
    sheet['D1'] = "test_Accuracy"
    logfile = open(
        args.output + "/" + args.name +
        '/log_run.txt',
        'w+'
    )
    logfile.write(str(args))
    
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    loss_sum = 0
    best_eval_accuracy = 0.0
    early_stop = 0
    decay_count = 0

    # Load the optimizer paramters
    optim = torch.optim.Adam(net.parameters(), lr=args.lr_base)
    net.to(device)
    loss_fn = args.loss_fn
    eval_accuracies = []
    for epoch in range(0, args.max_epoch):

        time_start = time.time()

        for step, (
                id,
                x,
                y,
                z,
                x1,
                ans,
        ) in enumerate(train_loader):

            loss_tmp = 0
            net.train()
            net.zero_grad()
            optim.zero_grad()


            x = x.to(torch.float32).cuda()
            y = y.to(torch.float32).cuda()
            z = z.to(torch.float32).cuda()
            x1 = x1.to(torch.float32).cuda()
            if args.task=='sentiment':
                ans = ans.to(torch.int64).to(x.device)
            else:
                ans = ans.cuda()

            pred,cl_loss = net(x, y, z, x1)
            #pred = pred.to(ans.dtype).to(x.device)
            #print(type(pred),pred.dtype)
            #print(type(ans),ans.dtype)
            
            loss = loss_fn(pred, ans)
            loss.backward()

            loss_sum += loss.cpu().data.numpy()
            loss_tmp += loss.cpu().data.numpy()

            print("\r[Epoch %2d][Step %4d/%4d] Loss: %.4f, Lr: %.2e, %4d m "
                  "remaining" % (
                      epoch + 1,
                      step,
                      int(len(train_loader.dataset) / args.batch_size),
                      loss_tmp / args.batch_size,
                      *[group['lr'] for group in optim.param_groups],
                      ((time.time() - time_start) / (step + 1)) * ((len(train_loader.dataset) / args.batch_size) - step) / 60,
                  ), end='          ')

            # Gradient norm clipping
            if args.grad_norm_clip > 0:
                nn.utils.clip_grad_norm_(
                    net.parameters(),
                    args.grad_norm_clip
                )

            optim.step()

        time_end = time.time()
        elapse_time = time_end-time_start
        print('Finished in {}s'.format(int(elapse_time)))
        accuracy, _ = evaluate(net, train_loader, args)
        sheet.cell(row=epoch + 1, column=1, value=epoch)
        sheet.cell(row=epoch + 1, column=2, value=accuracy)
        sheet.cell(row=epoch + 1, column=3, value=loss_sum / len(train_loader.dataset))
        
        epoch_finish = epoch + 1
        # Logging
        logfile.write(
            'Epoch: ' + str(epoch_finish) +
            ', Loss: ' + str(loss_sum / len(train_loader.dataset)) +
            ', Lr: ' + str([group['lr'] for group in optim.param_groups]) + '\n' +
            'Elapsed time: ' + str(int(elapse_time)) +
            ', Speed(s/batch): ' + str(elapse_time / step) +
            '\n\n'
        )

        # Eval
        if epoch_finish >= args.eval_start:
            print('Evaluation...')
            accuracy, _ = evaluate(net, eval_loader, args)
            sheet.cell(row=epoch + 1, column=4, value=accuracy)
            print('Accuracy :'+str(accuracy))
            eval_accuracies.append(accuracy)
            if accuracy > best_eval_accuracy:
                # Best
                state = {
                    'state_dict': net.state_dict(),
                    'optimizer': optim.state_dict(),
                    'args': args,
                }
                torch.save(
                    state,
                    args.output + "/" + args.name +
                    '/best'+str(args.seed)+'.pkl'
                )
                best_eval_accuracy = accuracy
                early_stop = 0

            elif decay_count < args.lr_decay_times:
                # Decay
                print('LR Decay...')
                decay_count += 1
                net.load_state_dict(torch.load(args.output + "/" + args.name +
                                               '/best'+str(args.seed)+'.pkl')['state_dict'])
                # adjust_lr(optim, args.lr_decay)
                for group in optim.param_groups:
                    group['lr'] *= args.lr_decay

            else:
                # Early stop
                early_stop += 1
                if early_stop == args.early_stop:
                    logfile.write('Early stop reached' + '\n')
                    print('Early stop reached')
                    logfile.write('best_overall_acc :' + str(best_eval_accuracy) + '\n\n')
                    print('best_eval_acc :' + str(best_eval_accuracy) + '\n\n')
                    os.rename(args.output + "/" + args.name +
                              '/best'+str(args.seed)+'.pkl',
                              args.output + "/" + args.name +
                              '/best' + str(best_eval_accuracy) + "_" + str(args.seed) + '.pkl')
                    logfile.close()
                    workbook.save("epoch_accuracy.xlsx")
                    return eval_accuracies
        gc.collect()
        torch.cuda.empty_cache()


        loss_sum = 0
    logfile.write('Not Early stop reached' + '\n')
    print('Max Epochs reached')
    logfile.write('best_overall_acc :' + str(best_eval_accuracy) + '\n\n')
    print('best_eval_acc :' + str(best_eval_accuracy) + '\n\n')
    os.rename(args.output + "/" + args.name +
              '/best'+str(args.seed)+'.pkl',
              args.output + "/" + args.name +
              '/best' + str(best_eval_accuracy) + "_" + str(args.seed) + '.pkl')
    logfile.close()
    workbook.save("epoch_accuracy.xlsx")


def evaluate(net, eval_loader, args):
    accuracy = []
    # device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    # print(device)
    # net.to(device)
    net.train(False)
    with torch.no_grad():
        preds = {}
        for step, (
                ids,
                x,
                y,
                z,
                x1,
                ans,
        ) in enumerate(eval_loader):
            x = x.to(torch.float32).cuda()
            y = y.to(torch.float32).cuda()
            z = z.to(torch.float32).cuda()
            x1 = x1.to(torch.float32).cuda()
            pred, cl_loss = net(x, y, z,x1)
            ans = ans.cuda()
            # if not eval_loader.dataset.private_set:
            #     ans = ans.cpu().data.numpy()
            pred_re = eval(args.pred_func)(pred)
            compare = torch.eq(pred_re, ans)
            accuracy += compare.cpu().numpy().tolist()
                
            # Save preds
            for id, p in zip(ids, pred.cpu().numpy()):
                preds[id] = p
    torch.cuda.empty_cache()
    gc.collect()
    net.train(True)
    return 100*np.mean(np.array(accuracy)), preds