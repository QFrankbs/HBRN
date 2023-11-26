import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pickle

import matplotlib.pyplot as plt

f4 = open(r"F:\New_retnet_multiModal\data\MOSEI\valid_audio.p",'rb')
f5 = open(r"F:\New_retnet_multiModal\data\MOSEI\valid_video.p",'rb')
f6 = open(r"F:\New_retnet_multiModal\data\MOSEI\valid_sentiment.p",'rb')
# f6 = open(r"E:\Semesterone\retention-multimodal\data\MOSEI\new_mosei\new_test_sentiment.p",'rb')
# f7 = open(r"E:\Semesterone\retention-multimodal\data\MOSEI\new_mosei\new_test_labels.p",'rb')

# data2 = pickle.load(f2)
# T = data2['-5B0PQx3Ep3k[0]']
# X = list(data2.values())
# x = X[0]
# data3 = pickle.load(f3)
data4 = pickle.load(f4)
data5 = pickle.load(f5)
data6 = pickle.load(f6)


# 提取特征数据并创建 DataFrame
features_data = []
y_values = []
for sample_name, features_array in data5.items():
    # 提取第一条所在行不为0的特征
    selected_features = features_array[np.any(features_array != 0, axis=1)]
    mean_features = np.mean(selected_features,axis = 0)
    mean_features = np.array(mean_features)
    new_mean_features = np.append(mean_features, data6[sample_name][0])
    features_data.append(new_mean_features)
    y_values.append(sample_name)
    # break

# for sample_name, features_array in data5.items():
#     # 提取第一条所在行不为0的特征
#     selected_features = features_array[np.any(features_array != 0, axis=1)]
#     for i in range(selected_features.shape[0]):
#         features_data.append( np.append(selected_features[i], data6[sample_name][0]))
#         y_values.append(sample_name)
# for sample_name, features_array in data5.items():
#     # 提取第一条所在行不为0的特征
#     selected_features = features_array[np.any(features_array != 0, axis=1)][0]
#     print(selected_features.shape,data6[sample_name][0])
#     new_selected_features = np.append(selected_features, data6[sample_name][0])
#     features_data.append(new_selected_features)
#     y_values.append(sample_name)


df = pd.DataFrame(features_data, columns=[f'Feature_{i+1}' for i in range(features_data[0].shape[0])], index=y_values)
df_corr = df.corr()

threshold_high = 0.0618
threshold_low = -0.05
relevant_features = df_corr[(df_corr['Feature_714'] > threshold_high) | (df_corr['Feature_714'] < threshold_low)]
relevant_feature_names = relevant_features.index.tolist()

# 打印选取的相关特征的列名
print(len(relevant_feature_names))
origin_nochange_features = []
for i in range(len(relevant_feature_names)):
    origin_nochange_features.append(int(relevant_feature_names[i][8:])-1)
    
origin_nochange_features.pop()


features_data = []

features_data_num = []
for sample_name, features_array in data5.items():
    # 提取第一条所在行不为0的特征
    selected_features = features_array[np.any(features_array != 0, axis=1)]
    features_data_num.append(selected_features.shape[0])
    for i in range(selected_features.shape[0]):
        features_data.append(selected_features[i][12:68])
features_data = np.array(features_data)
x = sum(features_data_num)
import numpy as np
from sklearn.decomposition import PCA
pca = PCA(n_components=14)
from sklearn.preprocessing import MinMaxScaler

# 创建 MinMaxScaler 对象
scaler = MinMaxScaler()

# 对数据进行归一化
features_data0 = scaler.fit_transform(features_data)

# 将数据进行 PCA 降维
reduced_data = pca.fit_transform(features_data0)

# 打印降维后的数据形状
print(reduced_data.shape)


features_data1 = []

for sample_name, features_array in data5.items():
    # 提取第一条所在行不为0的特征
    selected_features = features_array[np.any(features_array != 0, axis=1)]
    for i in range(selected_features.shape[0]):
        features_data1.append(selected_features[i][298:366])
features_data1 = np.array(features_data1)


# 对数据进行归一化
features_data1 = scaler.fit_transform(features_data1)
pca = PCA(n_components=16)
# 将数据进行 PCA 降维
features_data10 = pca.fit_transform(features_data1)

# 将数据进行 PCA 降维
reduced_data1 = pca.fit_transform(features_data10)

# 打印降维后的数据形状
print(reduced_data1.shape)



features_data2 = []

for sample_name, features_array in data5.items():
    # 提取第一条所在行不为0的特征
    selected_features = features_array[np.any(features_array != 0, axis=1)]
    for i in range(selected_features.shape[0]):
        features_data2.append( np.append(selected_features[i], data6[sample_name][0]))
features_data2 = np.array(features_data2)


from sklearn.preprocessing import StandardScaler

# 创建 StandardScaler 对象
scaler = StandardScaler()

# 对数据进行标准化


features_data20 = features_data2[:, origin_nochange_features]
reduced_data2 = scaler.fit_transform(features_data20)
# 将提取的列合并为一个新的数组
final_data = np.concatenate([reduced_data2,reduced_data1,reduced_data], axis=1)

# fig, ax = plt.subplots(figsize=(14, 14))
# import seaborn as sns

# sns.heatmap(df, square=True, vmax=1, vmin=-1, center=0,annot=True,cmap="coolwarm")
# plt.show()
# plt.savefig('heatmap_corr.png')
# # 计算皮尔逊相关系数矩阵
# correlation_matrix = df.corr()

# # 创建一个新的 Excel 工作簿
# wb = Workbook()
# ws = wb.active

# # 将相关系数矩阵写入 Excel 表格
# for r_idx, row in enumerate(correlation_matrix.values):
#     for c_idx, value in enumerate(row):
#         cell = ws.cell(row=r_idx + 1, column=c_idx + 1)
#         cell.value = value

#         # 根据相关性设置单元格颜色
#         if value > 0:
#             # 正相关，颜色越红
#             cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
#         else:
#             # 负相关，颜色越蓝
#             cell.fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')

# # 保存 Excel 文件
# wb.save("correlation_results.xlsx")
